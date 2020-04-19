from openpyxl import load_workbook
from sys import argv

wb2 = load_workbook(argv[1])


cell_nr=0

print("cell","cell_id","time","event.type","ctl")

for sheet in ['2014-11-25_pos 1','2014-11-25_pos 3','2014-11-25_pos 4','2014-12-09_pos 1','2014-12-09_pos 4']:
	ws = wb2[sheet]

	delta_t = ws['B3'].value / 60. / 60.
	video_length = ws['B4'].value * delta_t

	j=0
	cell_id=0
	observed_until=0
	died_at=0

	def force_val(x):
		if x == None:
			return 0
		if x == "":
			return 0
		if x == "CTL divides":
			return 0
		return x

	for row in ws.iter_rows(): # it brings a new method: iter_rows()
		# print "A", row[0].value, row[1].value
		j += 1
		if j >= 8 :
			if row[1].value != None and row[1].value != "" and row[1].value != "CTL divides":
				if cell_id != 0 :
					if died_at > 0 :
						print( cell_nr, cell_id, died_at, 1, 0 )
					elif observed_until > 0 and observed_until < video_length :
						print( cell_nr, cell_id, observed_until, -1, 0 )
					else :
						print( cell_nr, cell_id, video_length, 0, 0 )
				cell_nr += 1
				cell_id = force_val(row[1].value)
				observed_until = force_val(row[0].value) * delta_t
				died_at = force_val(row[6].value) * delta_t
			i = 10
			while len(row)>i and force_val(row[i].value)>0:
				print( cell_nr, cell_id, force_val(row[i].value) * delta_t, 2,
                                    force_val(row[2].value) )
				i += 1
	if died_at > 0 :
		print( cell_nr, cell_id, died_at, 1, 0 )
	elif observed_until > 0 :
		print( cell_nr, cell_id, observed_until, -1, 0 )
	else :
		print( cell_nr, cell_id, video_length, 0, 0 )
